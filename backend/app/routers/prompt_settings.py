import time
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.prompt_config import PromptConfig
from app.models.user import User
from app.schemas.prompt import (
    PromptConfigCreate,
    PromptConfigResponse,
    PromptConfigUpdate,
    PromptTestRequest,
    PromptTestResponse,
)
from app.services.auth_service import get_current_user
from app.services.prompt_service import SafeDict, get_quick_actions

router = APIRouter(prefix="/api/prompts", tags=["prompts"])


def _to_response(p: PromptConfig, is_overridden: bool = False) -> dict:
    return {
        "id": str(p.id),
        "prompt_key": p.prompt_key,
        "prompt_text": p.prompt_text,
        "category": p.category,
        "pipeline_stage": p.pipeline_stage,
        "variables": p.variables,
        "is_active": p.is_active,
        "is_system": p.user_id is None,
        "is_overridden": is_overridden,
        "seed_version": p.seed_version,
        "display_name": p.display_name,
        "description": p.description,
        "icon_name": p.icon_name,
        "sort_order": p.sort_order,
        "created_at": p.created_at.isoformat() if p.created_at else None,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None,
    }


@router.get("")
async def list_prompts(
    category: str | None = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List all prompts — merged view with user overrides."""
    # Get system defaults
    sys_q = select(PromptConfig).where(PromptConfig.user_id.is_(None))
    if category:
        sys_q = sys_q.where(PromptConfig.category == category)
    sys_result = await db.execute(sys_q.order_by(PromptConfig.category, PromptConfig.sort_order))
    system_prompts = {p.prompt_key: p for p in sys_result.scalars().all()}

    # Get user overrides
    usr_q = select(PromptConfig).where(PromptConfig.user_id == current_user.id)
    if category:
        usr_q = usr_q.where(PromptConfig.category == category)
    usr_result = await db.execute(usr_q.order_by(PromptConfig.sort_order))
    user_prompts = {p.prompt_key: p for p in usr_result.scalars().all()}

    # Merge
    items = []
    seen = set()
    for key, sys_p in system_prompts.items():
        usr_p = user_prompts.get(key)
        if usr_p:
            items.append(_to_response(usr_p, is_overridden=True))
        else:
            items.append(_to_response(sys_p, is_overridden=False))
        seen.add(key)

    # Custom user prompts not in system defaults
    for key, usr_p in user_prompts.items():
        if key not in seen:
            items.append(_to_response(usr_p, is_overridden=True))

    return items


@router.get("/quick-actions/{category}")
async def list_quick_actions(
    category: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get active quick actions for a category (e.g., quick_action.planner)."""
    return await get_quick_actions(category, current_user.id, db)


@router.get("/{prompt_key}")
async def get_prompt(
    prompt_key: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get effective prompt for current user."""
    result = await db.execute(
        select(PromptConfig).where(
            PromptConfig.prompt_key == prompt_key,
            or_(PromptConfig.user_id == current_user.id, PromptConfig.user_id.is_(None)),
        ).order_by(PromptConfig.user_id.desc().nulls_last()).limit(1)
    )
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Prompt not found")
    has_override = p.user_id is not None
    return _to_response(p, is_overridden=has_override)


@router.post("")
async def create_prompt(
    body: PromptConfigCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Create a user override or custom prompt."""
    prompt_key = body.prompt_key or f"custom.{current_user.id.hex[:8]}.{uuid.uuid4().hex[:8]}"
    p = PromptConfig(
        id=uuid.uuid4(),
        user_id=current_user.id,
        prompt_key=prompt_key,
        prompt_text=body.prompt_text,
        category=body.category,
        variables=body.variables,
        is_active=True,
        seed_version=0,
        display_name=body.display_name,
        description=body.description,
        icon_name=body.icon_name,
        sort_order=99,
    )
    db.add(p)
    await db.flush()
    await db.refresh(p)
    return _to_response(p, is_overridden=True)


@router.put("/{prompt_key}")
async def update_prompt(
    prompt_key: str,
    body: PromptConfigUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update or create user override for a prompt."""
    # Check if user already has an override
    result = await db.execute(
        select(PromptConfig).where(
            PromptConfig.prompt_key == prompt_key,
            PromptConfig.user_id == current_user.id,
        )
    )
    user_prompt = result.scalar_one_or_none()

    if user_prompt:
        # Update existing override
        if body.prompt_text is not None:
            user_prompt.prompt_text = body.prompt_text
        if body.is_active is not None:
            user_prompt.is_active = body.is_active
        if body.display_name is not None:
            user_prompt.display_name = body.display_name
        if body.description is not None:
            user_prompt.description = body.description
        if body.icon_name is not None:
            user_prompt.icon_name = body.icon_name
        await db.flush()
        await db.refresh(user_prompt)
        return _to_response(user_prompt, is_overridden=True)
    else:
        # Create new override by copying from system default
        sys_result = await db.execute(
            select(PromptConfig).where(
                PromptConfig.prompt_key == prompt_key,
                PromptConfig.user_id.is_(None),
            )
        )
        sys_prompt = sys_result.scalar_one_or_none()
        if not sys_prompt:
            raise HTTPException(status_code=404, detail="System prompt not found")

        override = PromptConfig(
            id=uuid.uuid4(),
            user_id=current_user.id,
            prompt_key=prompt_key,
            prompt_text=body.prompt_text or sys_prompt.prompt_text,
            category=sys_prompt.category,
            variables=sys_prompt.variables,
            is_active=body.is_active if body.is_active is not None else True,
            seed_version=0,
            display_name=body.display_name or sys_prompt.display_name,
            description=body.description or sys_prompt.description,
            icon_name=body.icon_name or sys_prompt.icon_name,
            sort_order=sys_prompt.sort_order,
        )
        db.add(override)
        await db.flush()
        await db.refresh(override)
        return _to_response(override, is_overridden=True)


@router.delete("/{prompt_key}")
async def reset_prompt(
    prompt_key: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete user override, reverting to system default."""
    result = await db.execute(
        select(PromptConfig).where(
            PromptConfig.prompt_key == prompt_key,
            PromptConfig.user_id == current_user.id,
        )
    )
    user_prompt = result.scalar_one_or_none()
    if not user_prompt:
        raise HTTPException(status_code=404, detail="No user override found")
    await db.delete(user_prompt)
    return {"status": "reset", "prompt_key": prompt_key}


@router.post("/by-id/{prompt_id}/duplicate")
async def duplicate_prompt(
    prompt_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Duplicate a prompt."""
    p = (await db.execute(select(PromptConfig).where(PromptConfig.id == prompt_id))).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Prompt not found")
    copy = PromptConfig(
        user_id=current_user.id, prompt_key=f"custom.{uuid.uuid4().hex[:8]}",
        prompt_text=p.prompt_text, category=p.category, pipeline_stage=p.pipeline_stage,
        variables=p.variables, is_active=True, seed_version=0,
        display_name=f"Copy of {p.display_name}", description=p.description,
        icon_name=p.icon_name, sort_order=p.sort_order + 1,
    )
    db.add(copy)
    await db.flush()
    return _to_response(copy, is_overridden=True)


@router.post("/by-id/{prompt_id}/revert")
async def revert_prompt(
    prompt_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Revert user override to system default."""
    p = (await db.execute(
        select(PromptConfig).where(PromptConfig.id == prompt_id, PromptConfig.user_id == current_user.id)
    )).scalar_one_or_none()
    if not p:
        raise HTTPException(404, "User override not found")
    # Find system version
    sys_p = (await db.execute(
        select(PromptConfig).where(PromptConfig.prompt_key == p.prompt_key, PromptConfig.user_id.is_(None))
    )).scalar_one_or_none()
    if sys_p:
        await db.delete(p)
        return _to_response(sys_p, is_overridden=False)
    raise HTTPException(400, "No system default to revert to")


@router.get("/export/xlsx")
async def export_prompts(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export all prompts to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from datetime import datetime

    prompts = (await db.execute(
        select(PromptConfig).where(or_(PromptConfig.user_id == current_user.id, PromptConfig.user_id.is_(None)))
        .order_by(PromptConfig.category, PromptConfig.sort_order)
    )).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Prompts"

    headers = ["prompt_key", "display_name", "category", "pipeline_stage", "description", "prompt_text", "variables", "icon_name", "sort_order", "is_active"]
    hfill = PatternFill(start_color="00338D", end_color="00338D", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont

    for i, p in enumerate(prompts, 2):
        ws.cell(row=i, column=1, value=p.prompt_key)
        ws.cell(row=i, column=2, value=p.display_name)
        ws.cell(row=i, column=3, value=p.category)
        ws.cell(row=i, column=4, value=p.pipeline_stage or "global")
        ws.cell(row=i, column=5, value=p.description or "")
        ws.cell(row=i, column=6, value=p.prompt_text)
        vlist = p.variables if isinstance(p.variables, list) else []
        ws.cell(row=i, column=7, value=", ".join(vlist) if vlist else "")
        ws.cell(row=i, column=8, value=p.icon_name or "")
        ws.cell(row=i, column=9, value=p.sort_order)
        ws.cell(row=i, column=10, value="Yes" if p.is_active else "No")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 100

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Prompts_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ─── Import ──────────────────────────────────────────────────────────────────

VALID_CATEGORIES = {"planner", "writer", "designer", "export", "quick_action_plan", "quick_action_write",
                    "quick_action.planner", "quick_action.writer", "global", "input", "custom"}

@router.get("/import/template")
async def download_template():
    """Download a blank Excel template for prompt import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import StreamingResponse
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Prompts"
    headers = ["prompt_key", "display_name", "category", "pipeline_stage", "description", "prompt_text", "variables", "icon_name", "sort_order", "is_active"]
    hfill = PatternFill(start_color="00338D", end_color="00338D", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont

    ws.append(["custom_example", "Example Custom Prompt", "custom", "global", "Example prompt for reference", "This is an example prompt. Use {audience} and {tone} variables.", "audience, tone", "Zap", 100, "Yes"])
    ws.append(["qa_plan_custom", "Custom Quick Action", "quick_action_plan", "step2_plan", "Custom quick action", "Apply this custom transformation to the slide.", "", "Sparkles", 50, "Yes"])

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["E"].width = 80

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": 'attachment; filename="Prompt_Import_Template.xlsx"'})


@router.post("/import/preview")
async def preview_import(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Parse an Excel file and return a preview of what will be created/updated."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx and .xls files are supported")

    import openpyxl
    from io import BytesIO

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "Max file size is 5MB")

    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    # Parse headers
    header_row = [str(cell.value or "").strip().lower() for cell in ws[1]]

    results = []
    for row in ws.iter_rows(min_row=2):
        vals = {}
        for i, cell in enumerate(row):
            if i < len(header_row) and header_row[i]:
                vals[header_row[i]] = cell.value

        pk = str(vals.get("prompt_key", "") or "").strip()
        dn = str(vals.get("display_name", "") or "").strip()
        pt = str(vals.get("prompt_text", "") or "").strip()
        cat = str(vals.get("category", "") or "").strip()

        if not pk and not dn and not pt:
            continue  # Skip empty rows

        errors = []
        if not pk:
            errors.append("prompt_key is required")
        if not dn:
            errors.append("display_name is required")
        if not pt:
            errors.append("prompt_text is required")
        if cat and cat not in VALID_CATEGORIES:
            errors.append(f"Invalid category: {cat}")

        # Check existing
        existing = None
        if pk:
            existing = (await db.execute(
                select(PromptConfig).where(
                    PromptConfig.prompt_key == pk,
                    or_(PromptConfig.user_id == current_user.id, PromptConfig.user_id.is_(None)),
                ).order_by(PromptConfig.user_id.desc().nulls_last()).limit(1)
            )).scalar_one_or_none()

        if errors:
            action = "error"
        elif existing is None:
            action = "create"
        elif (existing.prompt_text or "").strip() != pt.strip():
            action = "update"
        else:
            action = "skip"

        results.append({
            "row_number": row[0].row,
            "prompt_key": pk,
            "display_name": dn,
            "category": cat or "custom",
            "action": action,
            "errors": errors,
            "existing_text": existing.prompt_text if existing and action == "update" else None,
            "new_text": pt,
            "data": {k: str(v) if v is not None else "" for k, v in vals.items()},
        })

    summary = {
        "total": len(results),
        "create": sum(1 for r in results if r["action"] == "create"),
        "update": sum(1 for r in results if r["action"] == "update"),
        "skip": sum(1 for r in results if r["action"] == "skip"),
        "error": sum(1 for r in results if r["action"] == "error"),
    }

    return {"summary": summary, "rows": results}


class ImportRow(BaseModel):
    prompt_key: str
    display_name: str
    prompt_text: str
    category: str = "custom"
    pipeline_stage: str = "global"
    description: str = ""
    variables: str = ""
    icon_name: str = ""
    sort_order: int = 99
    is_active: str = "Yes"


class ImportApplyRequest(BaseModel):
    rows: list[dict]


@router.post("/import/apply")
async def apply_import(
    body: ImportApplyRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Apply the confirmed import rows."""
    created = 0
    updated = 0

    for row_data in body.rows:
        pk = row_data.get("prompt_key", "").strip()
        if not pk:
            continue

        existing = (await db.execute(
            select(PromptConfig).where(PromptConfig.prompt_key == pk, PromptConfig.user_id == current_user.id)
        )).scalar_one_or_none()

        variables_raw = row_data.get("variables", "")
        variables = [v.strip() for v in variables_raw.split(",") if v.strip()] if isinstance(variables_raw, str) else []

        is_active_raw = row_data.get("is_active", "Yes")
        is_active = str(is_active_raw).lower() in ("yes", "true", "1")

        sort_order = 99
        try:
            sort_order = int(row_data.get("sort_order", 99))
        except (ValueError, TypeError):
            pass

        if existing:
            existing.display_name = row_data.get("display_name", existing.display_name)
            existing.prompt_text = row_data.get("prompt_text", existing.prompt_text)
            existing.category = row_data.get("category", existing.category) or existing.category
            existing.pipeline_stage = row_data.get("pipeline_stage", existing.pipeline_stage) or existing.pipeline_stage
            existing.description = row_data.get("description") or existing.description
            existing.variables = variables or existing.variables
            existing.icon_name = row_data.get("icon_name") or existing.icon_name
            existing.sort_order = sort_order
            existing.is_active = is_active
            updated += 1
        else:
            new_p = PromptConfig(
                user_id=current_user.id, prompt_key=pk,
                display_name=row_data.get("display_name", pk),
                prompt_text=row_data.get("prompt_text", ""),
                category=row_data.get("category", "custom"),
                pipeline_stage=row_data.get("pipeline_stage", "global"),
                description=row_data.get("description"),
                variables=variables, icon_name=row_data.get("icon_name", "Zap"),
                sort_order=sort_order, is_active=is_active, seed_version=0,
            )
            db.add(new_p)
            created += 1

    await db.flush()
    return {"created": created, "updated": updated}


@router.post("/test")
async def test_prompt(
    body: PromptTestRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Test a prompt with variable substitution and optional LLM execution."""
    rendered = body.prompt_text.format_map(SafeDict(body.variables))
    resp = PromptTestResponse(rendered_text=rendered)

    if body.run_llm:
        try:
            from app.llm.factory import get_provider
            provider = get_provider(
                provider_name=body.llm_provider or "ollama",
                model=body.llm_model,
            )
            start = time.monotonic()
            result = await provider.generate(
                system_prompt="You are a test assistant.",
                user_prompt=rendered,
                json_mode=False,
            )
            latency = (time.monotonic() - start) * 1000
            resp.llm_response = result.get("text", str(result))
            resp.latency_ms = round(latency, 1)
        except Exception as e:
            resp.llm_response = f"Error: {e}"

    return resp
