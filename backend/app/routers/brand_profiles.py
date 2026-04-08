"""Brand profile CRUD API."""

import os
import re
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.brand_profile import BrandProfile
from app.models.user import User
from app.services.auth_service import get_current_user

router = APIRouter(prefix="/api/brand-profiles", tags=["brand-profiles"])

UPLOAD_DIR = "/app/uploads/brands"
DEFAULT_CHART_COLORS = ["#00338D", "#0091DA", "#483698", "#00A3A1", "#C6007E", "#FF6D00", "#009A44", "#6D2077"]

# All columns we want to serialize — avoids lazy-load issues with timestamps
_FIELDS = [
    "id", "user_id", "name", "description", "logo_path", "logo_position", "logo_size",
    "primary_color", "secondary_color", "accent_color", "background_color",
    "text_color", "text_secondary_color", "chart_colors",
    "font_heading", "font_body", "font_size_title", "font_size_subtitle",
    "font_size_body", "font_size_caption",
    "slide_header", "slide_footer", "slide_accent_line",
    "slide_background_style", "slide_gradient",
    "table_header_color", "table_header_text_color", "table_alternate_row",
    "table_alternate_color", "table_border_color", "table_style",
    "chart_style", "chart_show_grid", "chart_show_legend",
    "chart_legend_position", "chart_bar_radius",
    "is_default", "is_system",
]


def _to_dict(bp: BrandProfile) -> dict:
    d = {}
    for f in _FIELDS:
        v = getattr(bp, f)
        d[f] = str(v) if isinstance(v, uuid.UUID) else v
    d["logo_url"] = f"/uploads/brands/{bp.id}/logo{_logo_ext(bp.logo_path)}" if bp.logo_path else None
    d["chart_colors"] = bp.chart_colors or DEFAULT_CHART_COLORS
    return d


def _logo_ext(path: str | None) -> str:
    if not path:
        return ".png"
    return os.path.splitext(path)[1] or ".png"


# ── Schemas ───────────────────────────────────────────────────
class BrandCreate(BaseModel):
    name: str
    description: str | None = None
    primary_color: str = "#00338D"
    secondary_color: str = "#0091DA"
    accent_color: str = "#483698"
    background_color: str = "#FFFFFF"
    text_color: str = "#1A1A2E"
    text_secondary_color: str = "#6B7280"
    chart_colors: list[str] | None = None
    font_heading: str = "Arial"
    font_body: str = "Arial"
    font_size_title: int = 28
    font_size_subtitle: int = 18
    font_size_body: int = 14
    font_size_caption: int = 10


class BrandUpdate(BaseModel):
    name: str | None = None
    description: str | None = None
    logo_position: str | None = None
    logo_size: str | None = None
    primary_color: str | None = None
    secondary_color: str | None = None
    accent_color: str | None = None
    background_color: str | None = None
    text_color: str | None = None
    text_secondary_color: str | None = None
    chart_colors: list[str] | None = None
    font_heading: str | None = None
    font_body: str | None = None
    font_size_title: int | None = None
    font_size_subtitle: int | None = None
    font_size_body: int | None = None
    font_size_caption: int | None = None
    slide_header: dict | None = None
    slide_footer: dict | None = None
    slide_accent_line: dict | None = None
    slide_background_style: str | None = None
    slide_gradient: dict | None = None
    table_header_color: str | None = None
    table_header_text_color: str | None = None
    table_alternate_row: bool | None = None
    table_alternate_color: str | None = None
    table_border_color: str | None = None
    table_style: str | None = None
    chart_style: str | None = None
    chart_show_grid: bool | None = None
    chart_show_legend: bool | None = None
    chart_legend_position: str | None = None
    chart_bar_radius: int | None = None


# ── Endpoints ─────────────────────────────────────────────────
@router.get("")
async def list_profiles(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(BrandProfile).where(
            or_(BrandProfile.user_id == current_user.id, BrandProfile.is_system == True)  # noqa
        ).order_by(BrandProfile.is_system.desc(), BrandProfile.is_default.desc(), BrandProfile.name)
    )
    return [_to_dict(bp) for bp in result.scalars().all()]


@router.post("")
async def create_profile(
    body: BrandCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    bp = BrandProfile(user_id=current_user.id, **body.model_dump())
    if bp.chart_colors is None:
        bp.chart_colors = DEFAULT_CHART_COLORS
    db.add(bp)
    await db.flush()
    return _to_dict(bp)


@router.get("/{profile_id}")
async def get_profile(
    profile_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    bp = (await db.execute(select(BrandProfile).where(BrandProfile.id == profile_id))).scalar_one_or_none()
    if not bp:
        raise HTTPException(404, "Brand profile not found")
    return _to_dict(bp)


@router.put("/{profile_id}")
async def update_profile(
    profile_id: uuid.UUID,
    body: BrandUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    bp = (await db.execute(
        select(BrandProfile).where(BrandProfile.id == profile_id, BrandProfile.user_id == current_user.id)
    )).scalar_one_or_none()
    if not bp:
        # Check if system profile
        bp = (await db.execute(select(BrandProfile).where(BrandProfile.id == profile_id, BrandProfile.is_system == True))).scalar_one_or_none()  # noqa
        if bp:
            raise HTTPException(403, "Cannot edit system profiles")
        raise HTTPException(404, "Brand profile not found")

    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(bp, field, value)
    await db.flush()
    return _to_dict(bp)


@router.delete("/{profile_id}")
async def delete_profile(
    profile_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import or_
    bp = (await db.execute(
        select(BrandProfile).where(
            BrandProfile.id == profile_id,
            or_(BrandProfile.user_id == current_user.id, BrandProfile.is_system == True),  # noqa: E712
        )
    )).scalar_one_or_none()
    if not bp:
        raise HTTPException(404)
    # Clean up logo
    logo_dir = os.path.join(UPLOAD_DIR, str(bp.id))
    if os.path.exists(logo_dir):
        import shutil
        shutil.rmtree(logo_dir, ignore_errors=True)
    await db.delete(bp)
    return {"deleted": True}


@router.post("/{profile_id}/logo")
async def upload_logo(
    profile_id: uuid.UUID,
    file: UploadFile,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    bp = (await db.execute(select(BrandProfile).where(BrandProfile.id == profile_id))).scalar_one_or_none()
    if not bp:
        raise HTTPException(404)

    if file.content_type not in ("image/png", "image/jpeg", "image/svg+xml"):
        raise HTTPException(400, "Only PNG, JPG, SVG")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "Max 5MB")

    ext = {"image/png": ".png", "image/jpeg": ".jpg", "image/svg+xml": ".svg"}.get(file.content_type, ".png")
    save_dir = os.path.join(UPLOAD_DIR, str(profile_id))
    os.makedirs(save_dir, exist_ok=True)
    save_path = os.path.join(save_dir, f"logo{ext}")
    with open(save_path, "wb") as f:
        f.write(content)

    bp.logo_path = save_path
    await db.flush()
    return {"logo_url": f"/uploads/brands/{profile_id}/logo{ext}", "logo_path": save_path}


@router.delete("/{profile_id}/logo")
async def delete_logo(
    profile_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    bp = (await db.execute(select(BrandProfile).where(BrandProfile.id == profile_id))).scalar_one_or_none()
    if not bp:
        raise HTTPException(404)
    if bp.logo_path and os.path.exists(bp.logo_path):
        os.remove(bp.logo_path)
    bp.logo_path = None
    await db.flush()
    return {"deleted": True}


@router.post("/{profile_id}/set-default")
async def set_default(
    profile_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    # Unset all defaults for this user
    result = await db.execute(
        select(BrandProfile).where(BrandProfile.user_id == current_user.id, BrandProfile.is_default == True)  # noqa
    )
    for bp in result.scalars().all():
        bp.is_default = False

    bp = (await db.execute(select(BrandProfile).where(BrandProfile.id == profile_id))).scalar_one_or_none()
    if not bp:
        raise HTTPException(404)
    bp.is_default = True
    await db.flush()
    return {"is_default": True}


@router.post("/{profile_id}/duplicate")
async def duplicate_profile(
    profile_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    bp = (await db.execute(select(BrandProfile).where(BrandProfile.id == profile_id))).scalar_one_or_none()
    if not bp:
        raise HTTPException(404)

    new = BrandProfile(user_id=current_user.id, name=f"Copy of {bp.name}", is_system=False, is_default=False)
    for f in _FIELDS:
        if f in ("id", "user_id", "name", "is_default", "is_system"):
            continue
        setattr(new, f, getattr(bp, f))
    new.logo_path = None  # Don't copy logo file
    db.add(new)
    await db.flush()
    return _to_dict(new)


@router.post("/extract-colors")
async def extract_colors(
    file: UploadFile,
    current_user: User = Depends(get_current_user),
):
    import tempfile
    content = await file.read()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    tmp.write(content)
    tmp.close()
    try:
        from app.services.color_extractor import extract_colors_from_logo
        return extract_colors_from_logo(tmp.name)
    finally:
        os.unlink(tmp.name)


# ── Export / Import ────────────────────────────────────────────

_EXPORT_COLS = [
    "name", "description", "primary_color", "secondary_color", "accent_color",
    "background_color", "text_color", "text_secondary_color", "chart_colors",
    "font_heading", "font_body", "font_size_title", "font_size_subtitle",
    "font_size_body", "font_size_caption",
    "slide_header", "slide_footer", "slide_accent_line",
    "slide_background_style", "slide_gradient",
    "table_header_color", "table_header_text_color", "table_alternate_row",
    "table_alternate_color", "table_border_color", "table_style",
    "chart_style", "chart_show_grid", "chart_show_legend",
    "chart_legend_position", "chart_bar_radius",
    "logo_position", "logo_size", "is_default",
]


@router.get("/export/xlsx")
async def export_profiles(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export brand profiles to Excel with embedded logo images."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.drawing.image import Image as XlImage
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from datetime import datetime
    import json as json_mod

    profiles = (await db.execute(
        select(BrandProfile).where(
            or_(BrandProfile.user_id == current_user.id, BrandProfile.is_system == True)  # noqa
        ).order_by(BrandProfile.name)
    )).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Brand Profiles"

    headers = _EXPORT_COLS + ["logo_image"]
    hfill = PatternFill(start_color="00338D", end_color="00338D", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont

    logo_col_idx = len(headers)  # last column

    for i, bp in enumerate(profiles, 2):
        for j, col_name in enumerate(_EXPORT_COLS, 1):
            val = getattr(bp, col_name, None)
            # Serialize dicts/lists to JSON strings
            if isinstance(val, (dict, list)):
                val = json_mod.dumps(val, ensure_ascii=False)
            elif isinstance(val, bool):
                val = "Yes" if val else "No"
            ws.cell(row=i, column=j, value=val)

        # Embed logo image if exists
        if bp.logo_path and os.path.exists(bp.logo_path):
            try:
                img = XlImage(bp.logo_path)
                img.width = 60
                img.height = 60
                ws.add_image(img, f"{chr(64 + logo_col_idx)}{i}")
            except Exception:
                ws.cell(row=i, column=logo_col_idx, value="[logo file error]")
        else:
            ws.cell(row=i, column=logo_col_idx, value="")

    # Column widths
    ws.column_dimensions["A"].width = 30  # name
    ws.column_dimensions["B"].width = 40  # description
    for col_letter in ["C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 12

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Brand_Profiles_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/import/template")
async def download_brand_template():
    """Download a blank Excel template for brand profile import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import StreamingResponse
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Brand Profiles"
    headers = _EXPORT_COLS + ["logo_image"]
    hfill = PatternFill(start_color="00338D", end_color="00338D", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont

    # Example row
    ws.append([
        "My Brand", "Custom brand profile", "#00338D", "#0091DA", "#483698",
        "#FFFFFF", "#1A1A2E", "#6B7280", '["#00338D", "#0091DA", "#483698"]',
        "Arial", "Arial", 28, 18, 14, 10,
        "", "", "", "solid", "",
        "#00338D", "#FFFFFF", "Yes", "#F5F7FA", "#E5E7EB", "striped",
        "modern", "Yes", "Yes", "bottom", 4,
        "top-right", "medium", "No", "",
    ])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Brand_Profile_Import_Template.xlsx"'},
    )


@router.post("/import/preview")
async def preview_brand_import(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Parse an Excel file and return a preview of what will be created/updated."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx and .xls files are supported")

    import openpyxl
    from io import BytesIO

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(400, "Max file size is 10MB")

    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    header_row = [str(cell.value or "").strip().lower().replace(" ", "_") for cell in ws[1]]

    # Get existing profiles by name
    existing_result = await db.execute(
        select(BrandProfile).where(
            or_(BrandProfile.user_id == current_user.id, BrandProfile.is_system == True)  # noqa
        )
    )
    existing_by_name = {bp.name.lower(): bp for bp in existing_result.scalars().all()}

    results = []
    for row in ws.iter_rows(min_row=2):
        vals = {}
        for idx, cell in enumerate(row):
            if idx < len(header_row) and header_row[idx]:
                vals[header_row[idx]] = cell.value

        name = str(vals.get("name", "") or "").strip()
        if not name:
            continue  # skip empty rows

        errors = []
        if not name:
            errors.append("name is required")

        existing = existing_by_name.get(name.lower())

        if errors:
            action = "error"
        elif existing is None:
            action = "create"
        else:
            action = "update"

        results.append({
            "row_number": row[0].row,
            "name": name,
            "primary_color": str(vals.get("primary_color", "") or ""),
            "action": action,
            "errors": errors,
            "data": {k: str(v) if v is not None else "" for k, v in vals.items()},
        })

    summary = {
        "total": len(results),
        "create": sum(1 for r in results if r["action"] == "create"),
        "update": sum(1 for r in results if r["action"] == "update"),
        "skip": sum(1 for r in results if r["action"] == "skip"),
        "error": sum(1 for r in results if r["action"] == "error"),
    }

    return {"summary": summary, "rows": results}


class BrandImportApply(BaseModel):
    rows: list[dict]


@router.post("/import/apply")
async def apply_brand_import(
    body: BrandImportApply,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Apply the confirmed import rows."""
    import json as json_mod

    created = 0
    updated = 0

    for row_data in body.rows:
        name = str(row_data.get("name", "")).strip()
        if not name:
            continue

        # Find existing
        existing = (await db.execute(
            select(BrandProfile).where(
                BrandProfile.name == name,
                or_(BrandProfile.user_id == current_user.id, BrandProfile.user_id.is_(None)),
            ).order_by(BrandProfile.user_id.desc().nulls_last()).limit(1)
        )).scalar_one_or_none()

        # Build profile fields
        fields = {}
        for col in _EXPORT_COLS:
            if col in row_data and row_data[col] not in (None, ""):
                val = row_data[col]
                # Parse JSON fields
                if col in ("chart_colors", "slide_header", "slide_footer", "slide_accent_line", "slide_gradient"):
                    if isinstance(val, str):
                        try:
                            val = json_mod.loads(val)
                        except (json_mod.JSONDecodeError, ValueError):
                            pass
                # Parse boolean fields
                elif col in ("table_alternate_row", "chart_show_grid", "chart_show_legend", "is_default"):
                    val = str(val).lower() in ("yes", "true", "1")
                # Parse int fields
                elif col in ("font_size_title", "font_size_subtitle", "font_size_body", "font_size_caption", "chart_bar_radius"):
                    try:
                        val = int(val)
                    except (ValueError, TypeError):
                        continue
                fields[col] = val

        if existing:
            for k, v in fields.items():
                if k != "name":  # don't overwrite name
                    setattr(existing, k, v)
            updated += 1
        else:
            bp = BrandProfile(user_id=current_user.id, **fields)
            db.add(bp)
            created += 1

    await db.flush()
    return {"created": created, "updated": updated}


@router.post("/seed")
async def seed_profiles(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Seed system brand profiles. Only works if none exist."""
    existing = (await db.execute(select(BrandProfile).where(BrandProfile.is_system == True))).scalars().all()  # noqa
    if existing:
        return {"seeded": False, "message": "System profiles already exist"}

    presets = [
        {"name": "KPMG Official", "primary_color": "#00338D", "secondary_color": "#0091DA", "accent_color": "#483698",
         "chart_colors": ["#00338D", "#0091DA", "#483698", "#00A3A1", "#C6007E", "#FF6D00", "#009A44", "#6D2077"]},
        {"name": "Vision 2030", "primary_color": "#006C35", "secondary_color": "#FEFEFE", "accent_color": "#D4A843",
         "chart_colors": ["#006C35", "#D4A843", "#004D25", "#8B6914", "#2E8B57", "#DAA520", "#228B22", "#B8860B"]},
        {"name": "Clean Minimal", "primary_color": "#2563EB", "secondary_color": "#60A5FA", "accent_color": "#F59E0B",
         "chart_colors": ["#2563EB", "#60A5FA", "#F59E0B", "#10B981", "#EF4444", "#8B5CF6", "#EC4899", "#14B8A6"]},
        {"name": "Dark Executive", "primary_color": "#1E293B", "secondary_color": "#475569", "accent_color": "#F97316",
         "chart_colors": ["#1E293B", "#F97316", "#475569", "#0EA5E9", "#A855F7", "#EF4444", "#22C55E", "#EAB308"]},
        {
            "name": "SMO — Strategic Management Office",
            "description": "Official branding for مكتب الإدارة الاستراتيجية — Strategic Management Office, Kingdom of Saudi Arabia",
            "primary_color": "#1B6B4A", "secondary_color": "#2D8B6A", "accent_color": "#D4AF37",
            "background_color": "#FFFFFF", "text_color": "#1A1A1A", "text_secondary_color": "#6B7280",
            "chart_colors": ["#1B6B4A", "#2D8B6A", "#D4AF37", "#4AAB82", "#8B6914", "#3DA876", "#C49B2A", "#165A3E"],
            "font_heading": "Montserrat", "font_body": "Montserrat",
            "font_size_title": 28, "font_size_subtitle": 18, "font_size_body": 14, "font_size_caption": 10,
            "logo_position": "top-right", "logo_size": "medium",
            "slide_header": {"enabled": True, "color": "#1B6B4A", "height_percent": 7},
            "slide_footer": {"enabled": True, "color": "#1B6B4A", "height_percent": 4,
                             "show_page_number": True, "show_date": False,
                             "show_confidentiality": True, "confidentiality_text": "سري — Confidential"},
            "slide_accent_line": {"enabled": True, "position": "below_header", "thickness_px": 3, "color": "#D4AF37"},
            "slide_background_style": "solid",
            "table_header_color": "#1B6B4A", "table_header_text_color": "#FFFFFF",
            "table_alternate_row": True, "table_alternate_color": "#F0F7F4",
            "table_border_color": "#D1E7DD", "table_style": "striped",
            "chart_style": "modern", "chart_show_grid": True, "chart_show_legend": True,
            "chart_legend_position": "bottom", "chart_bar_radius": 4,
        },
    ]
    for p in presets:
        bp = BrandProfile(is_system=True, is_default=(p["name"] == "KPMG Official"), **p)
        db.add(bp)
    await db.flush()
    return {"seeded": True, "count": len(presets)}
